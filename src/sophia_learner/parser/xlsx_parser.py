"""
XLSX Parser Module

Provides parser for Microsoft Excel .xlsx files using openpyxl library
for text extraction and metadata retrieval with support for grouping/outline data.
"""

from pathlib import Path
from typing import Dict, Any, Optional, List, Tuple
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

from .base_parser import BaseParser, ParserError, EncryptedFileError
from ..security.sandbox import Sandbox


class XlsxParser(BaseParser):
    """
    Parser for Microsoft Excel .xlsx files.
    
    Uses openpyxl library with read_only=True for memory-efficient extraction
    of text content from worksheets. Includes support for grouping/outline data,
    formulas, and comprehensive metadata.
    
    Attributes:
        read_only: Whether to open workbook in read-only mode (memory efficient)
        data_only: Whether to get cell values or formulas (True = values)
        extract_formulas: Whether to extract formulas as text (default: False)
        extract_grouping: Whether to extract grouping/outline information (default: True)
        sandbox: Optional sandbox for resource-limited execution
    """
    
    def __init__(
        self, 
        sandbox: Optional[Sandbox] = None,
        read_only: bool = True,
        data_only: bool = True,
        extract_formulas: bool = False,
        extract_grouping: bool = True
    ):
        """
        Initialize the XLSX parser.
        
        Args:
            sandbox: Optional Sandbox instance for resource-limited execution
            read_only: Open workbook in read-only mode (default: True, memory efficient)
            data_only: Get cell values instead of formulas (default: True)
            extract_formulas: Whether to extract formula strings (default: False)
            extract_grouping: Whether to extract grouping/outline information (default: True)
        """
        super().__init__(sandbox)
        self.read_only = read_only
        self.data_only = data_only
        self.extract_formulas = extract_formulas
        self.extract_grouping = extract_grouping
    
    def _is_encrypted(self, file_path: Path) -> bool:
        """
        Check if the .xlsx file is encrypted/password-protected.
        
        Args:
            file_path: Path to the .xlsx file
            
        Returns:
            True if the file appears to be encrypted, False otherwise
        """
        try:
            # Try to open the workbook with minimal settings
            wb = load_workbook(
                str(file_path),
                read_only=True,
                data_only=True,
                keep_links=False
            )
            wb.close()
            return False
        except Exception as e:
            error_msg = str(e).lower()
            if 'encrypted' in error_msg or 'password' in error_msg:
                return True
            return False
    
    def _extract_grouping_info(self, sheet: Worksheet) -> Dict[str, Any]:
        """
        Extract grouping and outline information from a worksheet.
        
        Groups are hierarchical structures that allow collapsing/expanding
        rows or columns. This method extracts the grouping levels and structure.
        
        Args:
            sheet: openpyxl Worksheet object
            
        Returns:
            Dictionary containing grouping information:
            - row_groups: List of row groups with level, start, end, and visibility
            - col_groups: List of column groups with level, start, end, and visibility
            - outline_levels: Maximum outline levels for rows and columns
        """
        grouping_info = {
            'row_groups': [],
            'col_groups': [],
            'outline_levels': {'rows': 0, 'cols': 0}
        }
        
        if not self.extract_grouping:
            return grouping_info
        
        try:
            # Check for row grouping/outline
            if hasattr(sheet, 'row_dimensions'):
                row_groups = []
                current_group = None
                
                for row_idx in range(1, sheet.max_row + 1):
                    if row_idx in sheet.row_dimensions:
                        row_dim = sheet.row_dimensions[row_idx]
                        
                        # Check if row is part of a group
                        if hasattr(row_dim, 'outlineLevel') and row_dim.outlineLevel > 0:
                            level = row_dim.outlineLevel
                            grouping_info['outline_levels']['rows'] = max(
                                grouping_info['outline_levels']['rows'], 
                                level
                            )
                            
                            # Track visibility (collapsed/expanded)
                            is_collapsed = False
                            if hasattr(row_dim, 'hidden') and row_dim.hidden:
                                is_collapsed = True
                            
                            # Build group structure
                            if current_group is None or current_group['level'] != level:
                                if current_group:
                                    row_groups.append(current_group)
                                current_group = {
                                    'level': level,
                                    'start_row': row_idx,
                                    'end_row': row_idx,
                                    'collapsed': is_collapsed,
                                    'rows': [row_idx]
                                }
                            else:
                                current_group['end_row'] = row_idx
                                current_group['rows'].append(row_idx)
                                if is_collapsed:
                                    current_group['collapsed'] = True
                        else:
                            if current_group:
                                row_groups.append(current_group)
                                current_group = None
                
                if current_group:
                    row_groups.append(current_group)
                
                grouping_info['row_groups'] = row_groups
            
            # Check for column grouping/outline
            if hasattr(sheet, 'column_dimensions'):
                col_groups = []
                current_group = None
                
                for col_idx in range(1, sheet.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    if col_letter in sheet.column_dimensions:
                        col_dim = sheet.column_dimensions[col_letter]
                        
                        # Check if column is part of a group
                        if hasattr(col_dim, 'outlineLevel') and col_dim.outlineLevel > 0:
                            level = col_dim.outlineLevel
                            grouping_info['outline_levels']['cols'] = max(
                                grouping_info['outline_levels']['cols'], 
                                level
                            )
                            
                            # Track visibility
                            is_collapsed = False
                            if hasattr(col_dim, 'hidden') and col_dim.hidden:
                                is_collapsed = True
                            
                            # Build group structure
                            if current_group is None or current_group['level'] != level:
                                if current_group:
                                    col_groups.append(current_group)
                                current_group = {
                                    'level': level,
                                    'start_col': col_idx,
                                    'end_col': col_idx,
                                    'collapsed': is_collapsed,
                                    'columns': [col_idx]
                                }
                            else:
                                current_group['end_col'] = col_idx
                                current_group['columns'].append(col_idx)
                                if is_collapsed:
                                    current_group['collapsed'] = True
                        else:
                            if current_group:
                                col_groups.append(current_group)
                                current_group = None
                
                if current_group:
                    col_groups.append(current_group)
                
                grouping_info['col_groups'] = col_groups
        
        except Exception as e:
            self.logger.debug(f"Error extracting grouping info: {e}")
        
        return grouping_info
    
    def _extract_formulas_as_text(self, sheet: Worksheet, max_rows: int = None) -> List[Tuple[int, int, str]]:
        """
        Extract formulas from worksheet as text.
        
        Args:
            sheet: openpyxl Worksheet object
            max_rows: Maximum number of rows to process (optional)
            
        Returns:
            List of tuples (row, column, formula_string)
        """
        formulas = []
        
        if not self.extract_formulas:
            return formulas
        
        max_row = min(sheet.max_row, max_rows) if max_rows else sheet.max_row
        
        try:
            for row in range(1, max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row, col)
                    
                    # Check if cell has a formula
                    if hasattr(cell, 'formula') and cell.formula:
                        formulas.append((row, col, cell.formula))
                    
                    # Also check for array formulas
                    if hasattr(cell, 'array_formula') and cell.array_formula:
                        formulas.append((row, col, f"ARRAY:{cell.array_formula}"))
        
        except Exception as e:
            self.logger.debug(f"Error extracting formulas: {e}")
        
        return formulas
    
    def _cell_value_to_string(self, cell, include_formula: bool = False) -> str:
        """
        Convert a cell value to string, optionally including formula.
        
        Args:
            cell: openpyxl Cell object
            include_formula: Whether to include formula if extracting formulas
            
        Returns:
            String representation of the cell
        """
        if cell is None:
            return ""
        
        try:
            value = cell.value
            
            # Handle formula extraction if requested
            if include_formula and self.extract_formulas and hasattr(cell, 'formula') and cell.formula:
                formula = cell.formula
                if value is not None:
                    return f"{value} [FORMULA: {formula}]"
                else:
                    return f"[FORMULA: {formula}]"
            
            # Handle different value types
            if value is None:
                return ""
            elif isinstance(value, (int, float)):
                # Format numbers nicely
                if isinstance(value, float) and value.is_integer():
                    return str(int(value))
                return str(value)
            elif isinstance(value, datetime):
                return value.isoformat()
            elif isinstance(value, bool):
                return str(value)
            elif isinstance(value, str):
                return value.strip()
            elif isinstance(value, (list, tuple)):
                # Handle hyperlink or other complex types
                return str(value[0]) if value else ""
            else:
                return str(value)
        
        except Exception as e:
            self.logger.debug(f"Error converting cell to string: {e}")
            return ""
    
    def _extract_sheet_text(self, sheet: Worksheet, sheet_name: str) -> Tuple[str, Dict[str, Any]]:
        """
        Extract text and metadata from a single worksheet.
        
        Args:
            sheet: openpyxl Worksheet object
            sheet_name: Name of the worksheet
            
        Returns:
            Tuple of (sheet_text, sheet_metadata)
        """
        sheet_content = []
        sheet_metadata = {
            'name': sheet_name,
            'dimensions': None,
            'has_grouping': False,
            'row_count': 0,
            'column_count': 0,
            'cells_with_content': 0
        }
        
        # Get grouping information
        grouping_info = self._extract_grouping_info(sheet)
        if grouping_info['row_groups'] or grouping_info['col_groups']:
            sheet_metadata['has_grouping'] = True
            sheet_metadata['grouping'] = grouping_info
        
        # Get worksheet dimensions
        if sheet.max_row and sheet.max_column:
            sheet_metadata['dimensions'] = f"{sheet.max_row} rows × {sheet.max_column} columns"
            sheet_metadata['row_count'] = sheet.max_row
            sheet_metadata['column_count'] = sheet.max_column
        
        # Extract formulas if requested
        if self.extract_formulas:
            formulas = self._extract_formulas_as_text(sheet)
            if formulas:
                sheet_metadata['has_formulas'] = True
                sheet_metadata['formula_count'] = len(formulas)
        
        # Add sheet header with grouping info summary
        sheet_header = f"[SHEET: {sheet_name}]"
        if grouping_info['outline_levels']['rows'] > 0:
            sheet_header += f" (Row Groups: {grouping_info['outline_levels']['rows']} levels)"
        if grouping_info['outline_levels']['cols'] > 0:
            sheet_header += f" (Column Groups: {grouping_info['outline_levels']['cols']} levels)"
        
        sheet_content.append(sheet_header)
        
        # Track current row group level for indentation
        row_group_levels = {}
        if grouping_info['row_groups']:
            for group in grouping_info['row_groups']:
                for row in group['rows']:
                    row_group_levels[row] = group['level']
        
        # Process rows
        active_formulas = self.extract_formulas
        for row_idx in range(1, min(sheet.max_row + 1, 100000)):  # Safety limit
            row_cells = []
            
            # Get row prefix based on grouping level (for indentation)
            row_prefix = ""
            if row_idx in row_group_levels:
                level = row_group_levels[row_idx]
                row_prefix = "  " * level + f"[L{level}] "
            
            # Process columns
            has_content = False
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row_idx, col_idx)
                cell_value = self._cell_value_to_string(cell, active_formulas)
                
                if cell_value:
                    has_content = True
                    row_cells.append(cell_value)
            
            # If row has content, add it with appropriate prefix
            if has_content:
                row_text = row_prefix + '\t'.join(row_cells)
                sheet_content.append(row_text)
                sheet_metadata['cells_with_content'] += len(row_cells)
        
        # Add column grouping summary at the end
        if grouping_info['col_groups']:
            col_summary = ["[COLUMN GROUPS]"]
            for group in grouping_info['col_groups']:
                status = "collapsed" if group['collapsed'] else "expanded"
                col_summary.append(
                    f"  Level {group['level']}: columns {group['start_col']}-{group['end_col']} ({status})"
                )
            sheet_content.extend(col_summary)
        
        sheet_text = '\n'.join(sheet_content)
        return sheet_text, sheet_metadata
    
    def extract_text(self, file_path: Path) -> str:
        """
        Extract plain text from a .xlsx file.
        
        Processes each worksheet in the workbook, extracting cell values
        and preserving structure with tabs for columns and newlines for rows.
        Includes grouping/outline information if present.
        
        Args:
            file_path: Path to the .xlsx file
            
        Returns:
            Extracted plain text content with sheet structure preserved
            
        Raises:
            ParserError: If parsing fails
            EncryptedFileError: If the file is encrypted/password-protected
            FileNotFoundError: If file does not exist
            PermissionError: If file cannot be read
        """
        # Validate input
        self.validate_input(file_path)
        
        # Check for .xlsx extension
        if file_path.suffix.lower() != '.xlsx':
            self.logger.warning(f"File {file_path} does not have .xlsx extension")
        
        # Check if file is encrypted
        if self._is_encrypted(file_path):
            raise EncryptedFileError(
                f"File {file_path} appears to be encrypted or password-protected. "
                "This parser does not support encrypted .xlsx files."
            )
        
        self.logger.info(f"Extracting text from {file_path}")
        
        workbook = None
        try:
            # Open workbook with read_only and data_only settings
            workbook = load_workbook(
                str(file_path),
                read_only=self.read_only,
                data_only=self.data_only,
                keep_links=False  # Disable external links for security/speed
            )
            
            # Extract text from all sheets
            sheet_texts = []
            
            for sheet_name in workbook.sheetnames:
                self.logger.debug(f"Processing sheet: {sheet_name}")
                sheet = workbook[sheet_name]
                
                sheet_text, sheet_metadata = self._extract_sheet_text(sheet, sheet_name)
                if sheet_text:
                    sheet_texts.append(sheet_text)
                
                # Log grouping info if present
                if sheet_metadata.get('has_grouping'):
                    grouping = sheet_metadata.get('grouping', {})
                    self.logger.debug(
                        f"Sheet '{sheet_name}' has grouping: "
                        f"{len(grouping.get('row_groups', []))} row groups, "
                        f"{len(grouping.get('col_groups', []))} col groups"
                    )
            
            # Combine all sheets with double newlines
            full_text = '\n\n'.join(sheet_texts)
            
            # Sanitize output
            sanitized = self.sanitize_output(full_text)
            
            self.logger.debug(f"Extracted {len(sanitized)} characters from {file_path}")
            return sanitized
        
        except Exception as e:
            error_msg = str(e).lower()
            if 'encrypted' in error_msg or 'password' in error_msg:
                raise EncryptedFileError(f"File is encrypted: {file_path}")
            else:
                raise ParserError(f"Failed to parse .xlsx file {file_path}: {e}")
        
        finally:
            if workbook:
                workbook.close()
    
    def get_metadata(self, file_path: Path) -> Dict[str, Any]:
        """
        Extract metadata from a .xlsx file.
        
        Extracts workbook information including sheet names, sheet dimensions,
        properties, and grouping/outline statistics.
        
        Args:
            file_path: Path to the .xlsx file
            
        Returns:
            Dictionary containing metadata key-value pairs
            
        Raises:
            ParserError: If metadata extraction fails
            FileNotFoundError: If file does not exist
        """
        # Validate input
        self.validate_input(file_path)
        
        self.logger.info(f"Extracting metadata from {file_path}")
        
        workbook = None
        try:
            # Open workbook (read-only for metadata)
            workbook = load_workbook(
                str(file_path),
                read_only=True,
                data_only=True,
                keep_links=False
            )
            
            metadata = {
                'sheets': [],
                'sheet_count': len(workbook.sheetnames),
                'active_sheet': workbook.active.title if workbook.active else None,
                'has_macros': workbook.worksheets and hasattr(workbook, 'vba_archive'),
                'parser_settings': {
                    'read_only': self.read_only,
                    'data_only': self.data_only,
                    'extract_formulas': self.extract_formulas,
                    'extract_grouping': self.extract_grouping
                }
            }
            
            # Extract sheet information with grouping details
            total_rows = 0
            total_columns = 0
            total_cells_with_content = 0
            sheets_with_grouping = 0
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                sheet_info = {
                    'name': sheet_name,
                    'index': workbook.sheetnames.index(sheet_name),
                    'max_row': sheet.max_row,
                    'max_column': sheet.max_column,
                    'dimensions': f"{sheet.max_row}×{sheet.max_column}"
                }
                
                # Get grouping information for metadata
                if self.extract_grouping:
                    grouping_info = self._extract_grouping_info(sheet)
                    if grouping_info['row_groups'] or grouping_info['col_groups']:
                        sheets_with_grouping += 1
                        sheet_info['has_grouping'] = True
                        sheet_info['grouping'] = {
                            'row_group_levels': grouping_info['outline_levels']['rows'],
                            'col_group_levels': grouping_info['outline_levels']['cols'],
                            'row_group_count': len(grouping_info['row_groups']),
                            'col_group_count': len(grouping_info['col_groups'])
                        }
                    else:
                        sheet_info['has_grouping'] = False
                
                # Count rows and columns with content
                try:
                    rows_with_content = 0
                    for row in range(1, min(sheet.max_row + 1, 1000)):  # Sample for performance
                        for col in range(1, min(sheet.max_column + 1, 100)):
                            cell = sheet.cell(row, col)
                            if cell.value is not None:
                                rows_with_content += 1
                                break
                    sheet_info['rows_with_content_estimate'] = rows_with_content
                except Exception:
                    sheet_info['rows_with_content_estimate'] = 0
                
                metadata['sheets'].append(sheet_info)
                total_rows += sheet.max_row or 0
                total_columns = max(total_columns, sheet.max_column or 0)
            
            metadata['total_rows'] = total_rows
            metadata['max_columns'] = total_columns
            metadata['sheets_with_grouping'] = sheets_with_grouping
            
            # Extract workbook properties
            if hasattr(workbook, 'properties'):
                props = workbook.properties
                if props:
                    if props.title:
                        metadata['title'] = props.title
                    if props.subject:
                        metadata['subject'] = props.subject
                    if props.author:
                        metadata['author'] = props.author
                    if props.keywords:
                        metadata['keywords'] = props.keywords
                    if props.comments:
                        metadata['comments'] = props.comments
                    if props.created:
                        metadata['created'] = props.created.isoformat() if hasattr(props.created, 'isoformat') else str(props.created)
                    if props.modified:
                        metadata['modified'] = props.modified.isoformat() if hasattr(props.modified, 'isoformat') else str(props.modified)
                    if props.last_modified_by:
                        metadata['last_modified_by'] = props.last_modified_by
            
            # Add file information
            metadata['file_size_bytes'] = file_path.stat().st_size
            metadata['file_extension'] = file_path.suffix.lower()
            metadata['parser'] = 'openpyxl'
            
            self.logger.debug(f"Extracted {len(metadata)} metadata fields")
            return metadata
        
        except Exception as e:
            error_msg = str(e).lower()
            if 'encrypted' in error_msg or 'password' in error_msg:
                raise EncryptedFileError(f"File is encrypted: {file_path}")
            else:
                raise ParserError(f"Failed to extract metadata from {file_path}: {e}")
        
        finally:
            if workbook:
                workbook.close()
    
    def supports_encryption(self) -> bool:
        """
        Check if parser supports encrypted files.
        
        Override from BaseParser. .xlsx files may be encrypted with password
        protection. openpyxl cannot read encrypted files.
        
        Returns:
            False (encrypted .xlsx files are not supported)
        """
        return False
